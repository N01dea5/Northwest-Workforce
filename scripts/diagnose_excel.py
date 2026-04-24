"""Diagnostic: print a summary of what parse_macro_data sees and filters.

Usage:
    python3 scripts/diagnose_excel.py
    python3 scripts/diagnose_excel.py --excel path/to/file.xlsx

Outputs:
  - Unique Schedule Type values (with row counts)
  - Unique Client values (with row counts, classified or unmatched)
  - IsOnLocation value distribution
  - Per-month row counts at each filter stage
  - Per-month headcount (distinct workers) that would appear in the JSON
"""
from __future__ import annotations

import argparse
import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl required: pip install openpyxl", file=sys.stderr)
    raise

from dateutil.relativedelta import relativedelta

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_EXCEL = REPO_ROOT / "data" / "raw" / "Rapidcrews Macro Data.xlsx"

ONSITE_TYPES = {"Day Shift", "Night Shift", "RNR"}
CLIENT_ALIASES = [
    (re.compile(r"\bfmg\b|fortescue"), "Fortescue"),
    (re.compile(r"\brio\b|rtio|hamersley"), "Rio Tinto"),
    (re.compile(r"\bbhp\b|wa iron ore|\bwaio\b"), "BHP"),
    (re.compile(r"roy hill|hancock"), "Roy Hill/Hancock"),
]


def _norm(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def _classify_client(raw) -> str | None:
    key = _norm(raw).lower()
    if not key:
        return None
    for pat, name in CLIENT_ALIASES:
        if pat.search(key):
            return name
    return None


def _coerce_date(v) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _truthy(v) -> bool:
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(v)
    return str(v).strip().lower() in {"true", "yes", "y", "1", "t"}


def _pick_header(headers, *candidates):
    def key(s): return re.sub(r"[\s_\-]+", "", s.lower())
    index = {key(h): i for i, h in enumerate(headers)}
    for c in candidates:
        if key(c) in index:
            return index[key(c)]
    return None


def main(argv=None):
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    ap.add_argument("--current-month", default=None, help="YYYY-MM")
    args = ap.parse_args(argv)

    if not args.excel.exists():
        print(f"Excel not found: {args.excel}", file=sys.stderr)
        return 2

    if args.current_month:
        current = date(int(args.current_month[:4]), int(args.current_month[5:7]), 1)
    else:
        today = date.today()
        current = date(today.year, today.month, 1)

    window_start = current - relativedelta(months=3)
    window_end = current + relativedelta(months=4) - relativedelta(days=1)
    months = [
        (current - relativedelta(months=3) + relativedelta(months=i)).strftime("%Y-%m")
        for i in range(7)
    ]

    print(f"Current month : {current}")
    print(f"Window        : {window_start} → {window_end}")
    print(f"Months        : {months}\n")

    wb = openpyxl.load_workbook(args.excel, data_only=True, read_only=True)
    print(f"Sheets in workbook: {wb.sheetnames}\n")

    roster_sheet = "xpbi02 PersonnelRosterView"
    if roster_sheet not in wb.sheetnames:
        print(f"ERROR: sheet '{roster_sheet}' not found!")
        return 1

    ws = wb[roster_sheet]
    it = ws.iter_rows(values_only=True)
    header_row = next(it)
    headers = [_norm(c) for c in header_row]
    print(f"Roster sheet headers ({len(headers)} cols):")
    for i, h in enumerate(headers):
        print(f"  [{i:02d}] {h!r}")
    print()

    i_pid  = _pick_header(headers, "Personnel Id", "PersonnelId", "Personnel ID")
    i_dt   = _pick_header(headers, "Schedule Date", "ScheduleDate", "Date")
    i_type = _pick_header(headers, "Schedule Type", "ScheduleType", "Shift Type")
    i_cli  = _pick_header(headers, "Client", "Company")
    i_on   = _pick_header(headers, "IsOnLocation", "Is On Location", "On Location", "OnSite")

    print(f"Column index mapping:")
    print(f"  Personnel Id  → col {i_pid}  ({headers[i_pid] if i_pid is not None else 'NOT FOUND'})")
    print(f"  Schedule Date → col {i_dt}   ({headers[i_dt] if i_dt is not None else 'NOT FOUND'})")
    print(f"  Schedule Type → col {i_type} ({headers[i_type] if i_type is not None else 'NOT FOUND'})")
    print(f"  Client        → col {i_cli}  ({headers[i_cli] if i_cli is not None else 'NOT FOUND'})")
    print(f"  IsOnLocation  → col {i_on}   ({headers[i_on] if i_on is not None else 'not present'})")
    print()

    if None in (i_pid, i_dt, i_type, i_cli):
        print("ERROR: one or more required columns not found — cannot parse.")
        return 1

    stype_counter: Counter = Counter()
    client_counter: Counter = Counter()
    client_unmatched: Counter = Counter()
    on_loc_counter: Counter = Counter()

    # Per-month counts at each stage
    total_rows = 0
    after_window: Counter = Counter()
    after_stype: Counter = Counter()
    after_on_loc: Counter = Counter()
    after_client: Counter = Counter()
    workers_per_month: dict[str, set] = defaultdict(set)

    for row in it:
        if row is None:
            continue
        pid = _norm(row[i_pid])
        if not pid:
            continue
        total_rows += 1

        d = _coerce_date(row[i_dt])
        if d is None:
            continue
        mk = d.strftime("%Y-%m")
        if mk not in months:
            continue
        after_window[mk] += 1

        stype = _norm(row[i_type])
        stype_counter[stype] += 1
        if stype not in ONSITE_TYPES:
            continue
        after_stype[mk] += 1

        on_val = row[i_on] if i_on is not None else None
        on_loc_counter[str(on_val)] += 1
        if on_val is not None and not _truthy(on_val):
            continue
        after_on_loc[mk] += 1

        raw_client = _norm(row[i_cli])
        client = _classify_client(raw_client)
        client_counter[raw_client] += 1
        if client is None:
            client_unmatched[raw_client] += 1
            continue
        after_client[mk] += 1
        workers_per_month[mk].add(pid)

    print(f"Total non-empty rows read: {total_rows}\n")

    print("=== Schedule Type values (top 20) ===")
    for val, cnt in stype_counter.most_common(20):
        flag = "✓ ONSITE" if val in ONSITE_TYPES else "✗ filtered"
        print(f"  {cnt:6d}  {flag}  {val!r}")
    print()

    print("=== IsOnLocation values ===")
    for val, cnt in on_loc_counter.most_common():
        print(f"  {cnt:6d}  {val!r}")
    print()

    print("=== Client values (top 30) ===")
    for val, cnt in client_counter.most_common(30):
        classified = _classify_client(val)
        flag = f"→ {classified}" if classified else "✗ UNMATCHED"
        print(f"  {cnt:6d}  {flag:30s}  {val!r}")
    print()

    if client_unmatched:
        print("=== Unmatched client values (not one of the 4 NW majors) ===")
        for val, cnt in client_unmatched.most_common(30):
            print(f"  {cnt:6d}  {val!r}")
        print()

    print("=== Per-month row funnel ===")
    print(f"  {'Month':<10} {'In window':>10} {'After stype':>12} {'After OnLoc':>12} {'After client':>13} {'Workers':>8}")
    for m in months:
        print(f"  {m:<10} {after_window[m]:>10} {after_stype[m]:>12} {after_on_loc[m]:>12} {after_client[m]:>13} {len(workers_per_month[m]):>8}")
    print()


if __name__ == "__main__":
    raise SystemExit(main())
