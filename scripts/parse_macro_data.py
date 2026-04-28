"""Parse `Rapidcrews Macro Data.xlsx` into Northwest workforce JSON.

Consumes the same workbook the Southwest Shutdowns dashboard uses — drop the
same file into `data/raw/Rapidcrews Macro Data.xlsx` from Power Automate and
both dashboards stay in sync.

Input sheets used
-----------------
  xpbi02 PersonnelRosterView   daily roster (Personnel Id, Schedule Date,
                               Schedule Type, Client, Site, Job No, IsOnLocation)
  xll01 Personnel              worker master (Personnel Id, Given Names,
                               Surname, Primary Role, Hire Company,
                               optional Start Date)

Output
------
  data/workforce.json          canonical dashboard payload
  data/<client>.json           per-client pre-filtered payloads

Model
-----
* Current month = --current-month arg or today's month (first-of-month).
* Window = 3 months back + current + 3 months ahead (matches seed + dashboard).
* Hours per shift: Day Shift = 12h, Night Shift = 12h, RNR = 0h (rest day
  on-swing; the worker still counts as rostered).
* A row only contributes if Schedule Type is in ONSITE_TYPES and IsOnLocation
  is truthy (or missing, in which case on-site shift types are trusted).
* Each worker-month is assigned the CLIENT with the most rostered days that
  month (ties broken by name). Months with only gap/leave schedules are
  left client=null, hours=0.
* `committed=true` for any month strictly after the current month (the
  workbook is treated as authoritative for the full window).
* Top-20 positions = the 20 Primary Roles with the most rostered worker-months
  across the window. Workers whose role is outside the top 20 are folded into
  the JSON but won't appear in the summary table (same as the seed).
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

try:
    import openpyxl
except ImportError:
    print("openpyxl is required: pip install openpyxl", file=sys.stderr)
    raise

from dateutil.relativedelta import relativedelta

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_EXCEL = REPO_ROOT / "data" / "raw" / "Rapidcrews Macro Data.xlsx"
DEFAULT_OUT = REPO_ROOT / "data"

CLIENTS = ["Fortescue", "Rio Tinto", "BHP", "Roy Hill/Hancock"]

# Known aliases → canonical client name. Client values in the workbook are
# messy (they include site qualifiers, joint-venture names, etc.), so we
# match by substring on a normalised key.
CLIENT_ALIASES: list[tuple[re.Pattern[str], str]] = [
    (re.compile(r"\bfmg\b|fortescue"), "Fortescue"),
    (re.compile(r"\brio\b|rtio|hamersley"), "Rio Tinto"),
    (re.compile(r"\bbhp\b|wa iron ore|\bwaio\b"), "BHP"),
    (re.compile(r"roy hill|hancock"), "Roy Hill/Hancock"),
]

CLIENT_FILE_SLUG = {
    "Fortescue": "fortescue",
    "Rio Tinto": "rio-tinto",
    "BHP": "bhp",
    "Roy Hill/Hancock": "roy-hill",
}

ONSITE_TYPES = {"Day Shift", "Night Shift", "RNR"}
HOURS_PER_SHIFT = {"Day Shift": 12, "Night Shift": 12, "RNR": 0}

MONTHS_BEHIND = 3
MONTHS_AHEAD = 3
TOP_POSITIONS = 20

# ---------------------------------------------------------------------------
# small helpers
# ---------------------------------------------------------------------------

def _norm_text(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def _norm_name(v) -> str:
    return re.sub(r"[^a-z0-9]", "", _norm_text(v).lower())


def _classify_client(raw) -> str | None:
    key = _norm_text(raw).lower()
    if not key:
        return None
    for pat, canonical in CLIENT_ALIASES:
        if pat.search(key):
            return canonical
    return None


def _coerce_date(v) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _truthy(v) -> bool:
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(v)
    s = str(v).strip().lower()
    return s in {"true", "yes", "y", "1", "t"}


def _month_key(d: date) -> str:
    return d.strftime("%Y-%m")


def _month_start(d: date) -> date:
    return date(d.year, d.month, 1)


def _pick_header(headers: list[str], *candidates: str) -> int | None:
    """Locate the first matching column header, case-insensitive, whitespace-
    and punctuation-insensitive."""
    def key(s: str) -> str:
        return re.sub(r"[\s_\-]+", "", s.lower())
    index = {key(h): i for i, h in enumerate(headers)}
    for c in candidates:
        if key(c) in index:
            return index[key(c)]
    return None


# ---------------------------------------------------------------------------
# parser core
# ---------------------------------------------------------------------------

@dataclass
class RosterRow:
    personnel_id: str
    schedule_date: date
    schedule_type: str
    client: str
    job_no: str | None


@dataclass
class CalendarEntry:
    personnel_id: str
    start: date
    end: date
    client: str


@dataclass
class ShutdownAgg:
    shutdown_id: str
    shutdown_name: str
    client: str | None
    commence_date: date
    requested_by_trade: dict[str, set[str]] = field(default_factory=lambda: defaultdict(set))
    filled_by_trade: dict[str, set[str]] = field(default_factory=lambda: defaultdict(set))


@dataclass
class WorkerMaster:
    personnel_id: str
    name: str
    primary_role: str
    hire_company: str | None
    start_date: date | None
    discipline: str | None = None


def _read_headers(ws) -> tuple[list[str], Iterable]:
    it = ws.iter_rows(values_only=True)
    try:
        header_row = next(it)
    except StopIteration:
        return [], iter([])
    headers = [_norm_text(c) for c in header_row]
    return headers, it


def read_client_lookup(wb) -> dict[str, str]:
    """Build {client_id_guid: canonical_client_name} from xpbi02 ClientView.

    The PersonnelRosterView stores ClientId (GUID) in its Client column for
    historical rows.  We join via this lookup so alias matching works on the
    text name regardless of whether the roster cell holds a GUID or a name.
    """
    sheet = "xpbi02 ClientView"
    if sheet not in wb.sheetnames:
        return {}
    ws = wb[sheet]
    headers, it = _read_headers(ws)
    i_id   = _pick_header(headers, "ClientId", "Client Id", "Client ID", "Id")
    i_name = _pick_header(headers, "ClientName", "Client Name", "Name")
    if i_id is None or i_name is None:
        return {}
    lookup: dict[str, str] = {}
    for row in it:
        if row is None:
            continue
        cid  = _norm_text(row[i_id]).lower()
        name = _norm_text(row[i_name])
        if cid and name:
            lookup[cid] = name
    return lookup


def read_discipline_lookup(wb) -> dict[str, str]:
    """Return {trade_name_lower: discipline} from xpbi02 DisciplineTrade."""
    if "xpbi02 DisciplineTrade" not in wb.sheetnames:
        return {}
    ws = wb["xpbi02 DisciplineTrade"]
    headers, it = _read_headers(ws)
    i_trade = _pick_header(headers, "Trade")
    i_disc  = _pick_header(headers, "Discipline")
    if i_trade is None or i_disc is None:
        return {}
    out: dict[str, str] = {}
    for row in it:
        if row is None:
            continue
        trade = _norm_text(row[i_trade])
        disc  = _norm_text(row[i_disc])
        if trade and disc:
            out[trade.lower()] = disc
    return out


def _resolve_client(raw, client_lookup: dict[str, str]) -> str | None:
    """Classify a Client cell that may be a text name or a GUID."""
    text = _norm_text(raw)
    # Try direct alias match first (text name in cell).
    result = _classify_client(text)
    if result:
        return result
    # Fall back: treat cell value as a ClientId GUID and look up the name.
    resolved_name = client_lookup.get(text.lower())
    if resolved_name:
        return _classify_client(resolved_name)
    return None


def read_roster(wb, client_lookup: dict[str, str] | None = None) -> list[RosterRow]:
    if client_lookup is None:
        client_lookup = {}
    sheet_name = None
    for candidate in (
        "xpbi02 PersonnelRosterView",
        "xpbi02 DailyPersonnelSchedule",
        "PersonnelRosterView",
        "DailyPersonnelSchedule",
    ):
        if candidate in wb.sheetnames:
            sheet_name = candidate
            break
    if sheet_name is None:
        raise ValueError(
            "Workbook missing the roster sheet — expected one of "
            "'xpbi02 PersonnelRosterView' / 'xpbi02 DailyPersonnelSchedule'. "
            f"Have: {wb.sheetnames!r}"
        )
    ws = wb[sheet_name]
    headers, it = _read_headers(ws)
    # Old PersonnelRosterView layout has "Personnel Id"; new DailyPersonnelSchedule
    # uses "PersonnelId".  Try both.
    i_pid    = _pick_header(headers, "Personnel Id", "PersonnelId", "EmployeeId", "Employee Id")
    i_dt     = _pick_header(headers, "Schedule Date", "ScheduleDate", "ReportDate", "Report Date", "Date")
    # Schedule Type only exists on the old sheet; on the new layout hours are
    # credited per on-site day.
    i_type   = _pick_header(headers, "Schedule Type", "ScheduleType", "Shift Type")
    i_job    = _pick_header(headers, "Job No", "JobNo", "Job Number")
    i_on     = _pick_header(headers, "IsOnLocation", "Is On Location", "On Location", "OnSite")
    i_status = _pick_header(headers, "Status")
    i_active = _pick_header(headers, "employeeActive", "EmployeeActive", "Employee Active")

    i_client_name = _pick_header(headers, "Client Name", "ClientName")
    i_client_id   = _pick_header(headers, "ClientId", "Client Id", "Client ID")
    i_client_any  = _pick_header(headers, "Client", "Company")

    if i_pid is None or i_dt is None:
        raise ValueError(
            f"Roster sheet '{sheet_name}' missing Personnel/Employee Id or "
            f"Report/Schedule Date. Headers: {headers!r}"
        )
    if i_client_name is None and i_client_id is None and i_client_any is None:
        raise ValueError(f"Roster sheet '{sheet_name}' has no Client column. Headers: {headers!r}")

    def _client_from_row(row) -> str | None:
        if i_client_name is not None:
            r = _resolve_client(row[i_client_name], client_lookup)
            if r:
                return r
        if i_client_id is not None:
            r = _resolve_client(row[i_client_id], client_lookup)
            if r:
                return r
        if i_client_any is not None:
            return _resolve_client(row[i_client_any], client_lookup)
        return None

    # Deduplicate: one row per (personnel_id, date). The workbook emits both
    # 'onsite' and 'demobilised' rows for the same worker/day when a job ends
    # and another begins; both carry onsite=1 and would otherwise double-count
    # hours. Prefer 'onsite' status over 'demobilised'.
    STATUS_PRIORITY = {"onsite": 0, "demobilised": 1}
    seen: dict[tuple[str, date], tuple[int, RosterRow]] = {}

    for row in it:
        if row is None:
            continue
        pid = _norm_text(row[i_pid])
        if not pid:
            continue
        d = _coerce_date(row[i_dt])
        if d is None:
            continue
        status_str = _norm_text(row[i_status]).lower() if i_status is not None else ""
        if status_str in ("rejected", "declined", "late withdrawal"):
            continue
        if i_active is not None and row[i_active] is not None and not _truthy(row[i_active]):
            continue
        if i_on is not None and row[i_on] is not None and not _truthy(row[i_on]):
            # For future dates, only firm bookings (mobilising / confirmed) are
            # counted; soft statuses (contacted, planning, short list, etc.) are
            # excluded because they represent intent, not committed shifts.
            if not (d >= date.today() and status_str in ("mobilising", "confirmed")):
                continue

        if i_type is not None:
            stype = _norm_text(row[i_type])
            if stype not in ONSITE_TYPES:
                continue
        else:
            stype = "Day Shift"

        client = _client_from_row(row)
        if client is None:
            continue
        job_no = _norm_text(row[i_job]) if i_job is not None else None
        prio = STATUS_PRIORITY.get(status_str, 2)
        key = (pid, d)
        existing = seen.get(key)
        if existing is None or prio < existing[0]:
            seen[key] = (prio, RosterRow(pid, d, stype, client, job_no or None))

    return [r for _, r in seen.values()]


def read_personnel(
    wb,
    discipline_lookup: dict[str, str] | None = None,
) -> dict[str, WorkerMaster]:
    if "xll01 Personnel" not in wb.sheetnames:
        return {}
    discipline_lookup = discipline_lookup or {}
    ws = wb["xll01 Personnel"]
    headers, it = _read_headers(ws)
    i_pid = _pick_header(headers, "Personnel Id", "PersonnelId", "Personnel ID")
    i_given = _pick_header(headers, "Given Names", "First Name", "GivenNames")
    i_surname = _pick_header(headers, "Surname", "Last Name", "Family Name")
    i_full = _pick_header(headers, "Full Name", "Name")
    i_role = _pick_header(headers, "Primary Role", "PrimaryRole", "Role", "Trade")
    i_hire = _pick_header(headers, "Hire Company", "HireCompany", "Company")
    i_start = _pick_header(headers, "Start Date", "StartDate", "Hire Date", "Employment Start")

    out: dict[str, WorkerMaster] = {}
    for row in it:
        if row is None:
            continue
        pid = _norm_text(row[i_pid]) if i_pid is not None else ""
        if not pid:
            continue
        if i_full is not None and row[i_full]:
            name = _norm_text(row[i_full])
        else:
            given = _norm_text(row[i_given]) if i_given is not None else ""
            surname = _norm_text(row[i_surname]) if i_surname is not None else ""
            name = f"{given} {surname}".strip() or pid
        role = _norm_text(row[i_role]) if i_role is not None else ""
        hire = _norm_text(row[i_hire]) if i_hire is not None else ""
        start = _coerce_date(row[i_start]) if i_start is not None else None
        disc = discipline_lookup.get(role.lower())
        out[pid] = WorkerMaster(pid, name, role, hire or None, start, disc)
    return out


# ---------------------------------------------------------------------------
# aggregation
# ---------------------------------------------------------------------------

def build_months(current: date) -> list[date]:
    first = _month_start(current) - relativedelta(months=MONTHS_BEHIND)
    return [first + relativedelta(months=i) for i in range(MONTHS_BEHIND + 1 + MONTHS_AHEAD)]


@dataclass
class WorkerAgg:
    master: WorkerMaster
    # (month_key) -> (client -> {"days": int, "hours": int})
    buckets: dict[str, dict[str, dict]] = field(default_factory=lambda: defaultdict(lambda: defaultdict(lambda: {"days": 0, "hours": 0})))
    # Earliest schedule date seen; used as employment_start fallback.
    first_seen: date | None = None


MAX_CONSECUTIVE_DAYS = 13  # Day 14+ in an unbroken working run = fatigue stand-down


def aggregate(
    roster: list[RosterRow],
    personnel: dict[str, WorkerMaster],
    months: list[date],
    current_month: date,
) -> dict[str, WorkerAgg]:
    from calendar import monthrange

    start = months[0]
    end = months[-1] + relativedelta(months=1) - relativedelta(days=1)

    aggs: dict[str, WorkerAgg] = {}
    # Track distinct dates per (pid, month_key) for full-roster detection.
    dates_seen: dict[str, dict[str, set]] = defaultdict(lambda: defaultdict(set))
    # Preserve original coverage before stand-down adjustment so "booked every
    # day of month" halving still applies even when stand-down removes day 14s.
    raw_dates_seen: dict[str, dict[str, set]] = defaultdict(lambda: defaultdict(set))
    # Per-day detail for 13-day consecutive rule: {pid: {date: (client, hrs)}}
    day_detail: dict[str, dict[date, tuple[str, int]]] = defaultdict(dict)

    for r in roster:
        if r.schedule_date < start or r.schedule_date > end:
            continue
        master = personnel.get(r.personnel_id) or WorkerMaster(
            personnel_id=r.personnel_id,
            name=r.personnel_id,
            primary_role="Unknown",
            hire_company=None,
            start_date=None,
        )
        a = aggs.setdefault(r.personnel_id, WorkerAgg(master=master))
        if a.first_seen is None or r.schedule_date < a.first_seen:
            a.first_seen = r.schedule_date
        mk = _month_key(r.schedule_date)
        hrs = HOURS_PER_SHIFT.get(r.schedule_type, 0)
        dates_seen[r.personnel_id][mk].add(r.schedule_date)
        raw_dates_seen[r.personnel_id][mk].add(r.schedule_date)
        day_detail[r.personnel_id][r.schedule_date] = (r.client, hrs)
        bucket = a.buckets[mk][r.client]
        bucket["days"] += 1
        bucket["hours"] += hrs

    # Apply 13-day consecutive stand-down rule.
    # Scan each worker's working days (hours > 0) in date order. When a
    # consecutive run reaches 14 days, that 14th day is zeroed out as a
    # fatigue stand-down, which also breaks the run so day 15 starts fresh.
    from datetime import timedelta
    for pid, details in day_detail.items():
        if pid not in aggs:
            continue
        working_dates = sorted(d for d, (_, h) in details.items() if h > 0)
        run_len = 0
        prev_d: date | None = None
        for d in working_dates:
            if prev_d is not None and (d - prev_d).days == 1:
                run_len += 1
            else:
                run_len = 1
            if run_len > MAX_CONSECUTIVE_DAYS:
                # Stand-down: remove this day's hours from its bucket.
                client, hrs = details[d]
                mk = _month_key(d)
                bucket = aggs[pid].buckets.get(mk, {}).get(client, {})
                if bucket and hrs > 0:
                    bucket["hours"] = max(0, bucket["hours"] - hrs)
                    bucket["days"]  = max(0, bucket["days"] - 1)
                dates_seen[pid][mk].discard(d)
                # Break the run — next working day starts a new streak.
                run_len = 0
                prev_d = None
            else:
                prev_d = d

    # If a worker has an entry for every calendar day of a month, the
    # scheduling system is recording them as on-roster continuously (e.g. a
    # maintenance contract). Assume 50% actual on-site attendance and halve
    # the hours for that month.
    for pid, agg in aggs.items():
        for mk, by_client in agg.buckets.items():
            yr, mo = int(mk[:4]), int(mk[5:7])
            _, days_in_month = monthrange(yr, mo)
            if len(raw_dates_seen[pid][mk]) >= days_in_month:
                for b in by_client.values():
                    b["hours"] = round(b["hours"] / 2)
                    b["days"]  = round(b["days"] / 2)

    return aggs


def dominant_client_per_month(agg: WorkerAgg, month_key: str) -> tuple[str | None, int]:
    """Return (client, hours) for the month. Client is the one with most
    rostered days; hours is the sum across all clients that month (so workers
    who split between clients aren't double-counted)."""
    by_client = agg.buckets.get(month_key) or {}
    if not by_client:
        return None, 0
    client = max(by_client.items(), key=lambda kv: (kv[1]["days"], kv[0]))[0]
    hours = sum(v["hours"] for v in by_client.values())
    return client, hours


def pick_primary_client(agg: WorkerAgg) -> str:
    counter: Counter[str] = Counter()
    for month in agg.buckets.values():
        for client, v in month.items():
            counter[client] += v["days"]
    if not counter:
        return agg.master.hire_company and _classify_client(agg.master.hire_company) or "Fortescue"
    return counter.most_common(1)[0][0]


def pick_top_positions(aggs: dict[str, WorkerAgg]) -> list[str]:
    counter: Counter[str] = Counter()
    for a in aggs.values():
        role = a.master.primary_role or "Unknown"
        # Count distinct worker-months, not just workers, so higher-volume
        # positions bubble up.
        counter[role] += sum(1 for m in a.buckets.values() if m)
    return [role for role, _ in counter.most_common(TOP_POSITIONS)]


# ---------------------------------------------------------------------------
# payload assembly
# ---------------------------------------------------------------------------

def worker_payload(
    agg: WorkerAgg,
    months: list[date],
    current_month: date,
) -> dict:
    mk_months = [_month_key(m) for m in months]
    cur_key = _month_key(current_month)
    cur_idx = mk_months.index(cur_key)

    monthly: dict[str, dict] = {}
    for i, mk in enumerate(mk_months):
        client, hours = dominant_client_per_month(agg, mk)
        is_future = i > cur_idx
        monthly[mk] = {
            "client": client,
            "hours": hours,
            "committed": is_future,
        }

    primary_client = pick_primary_client(agg)
    employment_start = agg.master.start_date or agg.first_seen or months[0]

    return {
        "id": "w-" + _norm_name(agg.master.name)[:14],
        "name": agg.master.name,
        "position": agg.master.primary_role or "Unknown",
        "discipline": agg.master.discipline or "Other",
        "primary_client": primary_client,
        "employment_start": employment_start.isoformat(),
        "employment_end": None,
        "monthly": monthly,
    }


def read_calendar_view(wb, client_lookup: dict[str, str]) -> list[CalendarEntry]:
    """DEPRECATED: PersonnelCalendarView is an unavailability calendar
    (Medical / Training / Accom / Travel), not a work assignment log.
    Retained as a no-op for backward compatibility; real historical
    work data now comes from xpbipr DailyEmployeeSchedule.
    """
    return []


def read_daily_employee_schedule(wb, client_lookup: dict[str, str]) -> list[RosterRow]:
    """Read xpbipr DailyEmployeeSchedule — daily roster rows for all jobs
    including closed shutdowns.

    PersonnelRosterView is filtered to active jobs only, so historical rows
    for workers on completed shutdowns disappear there.  DailyEmployeeSchedule
    is the long-tail view that keeps them.

    Columns used: EmployeeId, ReportDate, Status, OnSite, Client / ClientId,
    EmployeeActive.  Schedule Type is absent so we credit 12 h per on-site day.
    """
    sheet_name = None
    for candidate in (
        "ACTIVE_SHUTDOWNS",
        "ACTIVE_SHUTDOWN",
        "xpbipr DailyEmployeeSchedule",
        "xpbi0r DailyEmployeeSchedule",
        "xpbi02 DailyEmployeeSchedule",
        "DailyEmployeeSchedule",
    ):
        if candidate in wb.sheetnames:
            sheet_name = candidate
            break
    if sheet_name is None:
        return []

    ws = wb[sheet_name]
    headers, it = _read_headers(ws)

    i_pid    = _pick_header(headers, "EmployeeId", "Employee Id", "Personnel Id", "PersonnelId")
    i_dt     = _pick_header(headers, "ReportDate", "Report Date", "Date", "Schedule Date")
    i_status = _pick_header(headers, "Status")
    i_onsite = _pick_header(headers, "OnSite", "On Site", "IsOnLocation")
    i_cname  = _pick_header(headers, "Client Name", "ClientName", "Client", "Company")
    i_cid    = _pick_header(headers, "ClientId", "Client Id", "Client ID")
    i_active = _pick_header(headers, "EmployeeActive", "Employee Active")

    if i_pid is None or i_dt is None:
        return []

    out: list[RosterRow] = []
    for row in it:
        if row is None:
            continue
        pid = _norm_text(row[i_pid])
        if not pid:
            continue
        d = _coerce_date(row[i_dt])
        if d is None:
            continue
        if i_status is not None:
            status = _norm_text(row[i_status]).lower()
            if status == "rejected":
                continue
        if i_active is not None and row[i_active] is not None and not _truthy(row[i_active]):
            continue
        if i_onsite is not None and row[i_onsite] is not None and not _truthy(row[i_onsite]):
            continue

        client = None
        for ci in (i_cname, i_cid):
            if ci is not None:
                client = _resolve_client(row[ci], client_lookup)
                if client:
                    break
        if client is None:
            continue

        out.append(RosterRow(
            personnel_id=pid,
            schedule_date=d,
            schedule_type="Day Shift",
            client=client,
            job_no=None,
        ))
    return out


def dedupe_ids(workers: list[dict]) -> list[dict]:
    seen: set[str] = set()
    for w in workers:
        base = w["id"]
        cand = base
        n = 1
        while cand in seen:
            n += 1
            cand = f"{base}{n}"
        seen.add(cand)
        w["id"] = cand
    return workers


def _sheet_or_none(wb, candidates: tuple[str, ...]):
    for c in candidates:
        if c in wb.sheetnames:
            return wb[c]
    return None


def read_shutdown_fulfilment(
    wb,
    client_lookup: dict[str, str],
) -> tuple[list[dict], dict[str, dict[str, str]]]:
    """Extract shutdown requested/filled counts and worker outcomes.

    Source: DailyPersonnelSchedule rows. Rows are daily, so counts are deduped
    to unique (shutdown, trade, worker) to represent positions.
    """
    ws = _sheet_or_none(
        wb,
        (
            "xpbi02 DailyPersonnelSchedule",
            "xpbipr DailyEmployeeSchedule",
            "xpbi0r DailyEmployeeSchedule",
            "xpbi02 DailyEmployeeSchedule",
            "DailyEmployeeSchedule",
        ),
    )
    if ws is None:
        return [], {}

    headers, it = _read_headers(ws)
    i_pid    = _pick_header(headers, "PersonnelId", "Personnel Id", "EmployeeId", "Employee Id")
    i_dt     = _pick_header(headers, "ReportDate", "Report Date", "Date", "Schedule Date")
    i_status = _pick_header(headers, "Status")
    i_trade  = _pick_header(headers, "Trade", "Primary Role", "Role")
    i_jid    = _pick_header(headers, "JobId", "Job No", "JobNo")
    i_jname  = _pick_header(headers, "JobGroup", "Job Group", "Shutdown", "Project")
    i_jstart = _pick_header(headers, "JobStart", "Start", "Start Date")
    i_client_name = _pick_header(headers, "Client", "Client Name", "ClientName")
    i_client_id   = _pick_header(headers, "ClientId", "Client Id", "Client ID")

    if i_dt is None or i_status is None or i_trade is None:
        return [], {}

    # Fulfilment tracking is based on job-planning pipeline states.
    REQUEST_STATUSES = {
        "contacted", "planning", "short list", "confirmed", "mobilising",
        "declined", "rejected", "late withdrawal",
    }
    FILLED_STATUSES = {"confirmed", "mobilising"}
    WORKED_STATUSES = {"onsite", "demobilised"}
    DECLINED_STATUSES = {"declined", "rejected", "late withdrawal"}
    TRACKED_STATUSES = REQUEST_STATUSES | WORKED_STATUSES

    by_shutdown: dict[str, ShutdownAgg] = {}
    worker_outcomes: dict[str, dict[str, str]] = defaultdict(dict)

    def _row_client(row):
        for idx in (i_client_name, i_client_id):
            if idx is None:
                continue
            c = _resolve_client(row[idx], client_lookup)
            if c:
                return c
        return None

    for row in it:
        if row is None:
            continue
        pid = _norm_text(row[i_pid]) if i_pid is not None else ""
        report_date = _coerce_date(row[i_dt])
        status = _norm_text(row[i_status]).lower()
        trade = _norm_text(row[i_trade]) or "Unknown"
        if report_date is None or not status or status not in TRACKED_STATUSES:
            continue

        jid = _norm_text(row[i_jid]) if i_jid is not None else ""
        jname = _norm_text(row[i_jname]) if i_jname is not None else ""
        jstart = _coerce_date(row[i_jstart]) if i_jstart is not None else None
        commence = jstart or report_date
        shutdown_id = f"{jid or jname or 'unknown'}-{commence.isoformat()}"
        shutdown_name = jname or (f"Job {jid}" if jid else f"Shutdown {commence.isoformat()}")
        client = _row_client(row)

        agg = by_shutdown.get(shutdown_id)
        if agg is None:
            agg = ShutdownAgg(
                shutdown_id=shutdown_id,
                shutdown_name=shutdown_name,
                client=client,
                commence_date=commence,
            )
            by_shutdown[shutdown_id] = agg
        if client and not agg.client:
            agg.client = client

        if pid and status in REQUEST_STATUSES:
            agg.requested_by_trade[trade].add(pid)
            if status in FILLED_STATUSES:
                agg.filled_by_trade[trade].add(pid)

        if pid:
            current = worker_outcomes[pid].get(shutdown_id)
            if status in WORKED_STATUSES:
                worker_outcomes[pid][shutdown_id] = "worked"
            elif status in DECLINED_STATUSES and current != "worked":
                worker_outcomes[pid][shutdown_id] = "declined"

    shutdowns = []
    for s in sorted(by_shutdown.values(), key=lambda x: (x.commence_date, x.shutdown_name)):
        all_trades = sorted(set(s.requested_by_trade) | set(s.filled_by_trade))
        trades = []
        req_total = 0
        fill_total = 0
        for t in all_trades:
            requested = len(s.requested_by_trade.get(t, set()))
            filled = len(s.filled_by_trade.get(t, set()))
            req_total += requested
            fill_total += filled
            trades.append({
                "trade": t,
                "requested": requested,
                "filled": filled,
                "gap": requested - filled,
                "fill_rate": (filled / requested) if requested else None,
            })
        shutdowns.append({
            "id": s.shutdown_id,
            "name": s.shutdown_name,
            "client": s.client,
            "commence_date": s.commence_date.isoformat(),
            "commence_month": _month_key(s.commence_date),
            "requested_total": req_total,
            "filled_total": fill_total,
            "gap_total": req_total - fill_total,
            "fill_rate": (fill_total / req_total) if req_total else None,
            "trades": trades,
        })
    return shutdowns, worker_outcomes


def build_payload(excel_path: Path, current_month: date) -> dict:
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    client_lookup = read_client_lookup(wb)
    discipline_lookup = read_discipline_lookup(wb)
    # PersonnelRosterView: active jobs only, but carries Schedule Type
    # detail (Day Shift / Night Shift / RNR) so gives accurate hours.
    roster = read_roster(wb, client_lookup)
    # DailyEmployeeSchedule: long-tail view including closed shutdowns.
    # Fills historical gaps the RosterView leaves behind.
    daily_extra = read_daily_employee_schedule(wb, client_lookup)
    # Dedupe: prefer roster rows for any (worker, date, client) key.
    seen = {(r.personnel_id, r.schedule_date, r.client) for r in roster}
    extra = [r for r in daily_extra if (r.personnel_id, r.schedule_date, r.client) not in seen]
    all_rows = roster + extra

    personnel = read_personnel(wb, discipline_lookup)
    months = build_months(current_month)
    aggs = aggregate(all_rows, personnel, months, current_month)
    shutdowns, worker_shutdown_outcomes = read_shutdown_fulfilment(wb, client_lookup)

    top_positions = pick_top_positions(aggs)
    # If the workbook lacks enough roles, pad with an empty placeholder so
    # the dashboard's top-20 table still renders.
    while len(top_positions) < TOP_POSITIONS:
        top_positions.append(f"(Unassigned {len(top_positions) + 1})")

    workers = []
    for pid, agg in aggs.items():
        w = worker_payload(agg, months, current_month)
        w["shutdown_outcomes"] = worker_shutdown_outcomes.get(pid, {})
        workers.append(w)
    workers = dedupe_ids(sorted(workers, key=lambda w: w["name"]))

    return {
        "generated_at": date.today().isoformat(),
        "current_month": _month_key(current_month),
        "reporting_months": [_month_key(m) for m in months],
        "months_behind": MONTHS_BEHIND,
        "months_ahead": MONTHS_AHEAD,
        "positions_top20": top_positions[:TOP_POSITIONS],
        "clients": CLIENTS,
        "workers": workers,
        "shutdowns": shutdowns,
    }


def write_json(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, sort_keys=False), encoding="utf-8")


def write_outputs(payload: dict, out_dir: Path) -> None:
    write_json(out_dir / "workforce.json", payload)
    for client in CLIENTS:
        slug = CLIENT_FILE_SLUG[client]
        filtered = [
            w for w in payload["workers"]
            if any(m.get("client") == client for m in w["monthly"].values())
        ]
        per_client = dict(payload)
        per_client["workers"] = filtered
        per_client["scoped_client"] = client
        write_json(out_dir / f"{slug}.json", per_client)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_current_month(s: str | None) -> date:
    if not s:
        return _month_start(date.today())
    return _month_start(datetime.strptime(s, "%Y-%m").date())


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    ap.add_argument("--out", type=Path, default=DEFAULT_OUT)
    ap.add_argument(
        "--current-month",
        default=None,
        help="YYYY-MM; defaults to this month.",
    )
    args = ap.parse_args(argv)

    if not args.excel.exists():
        print(f"Excel not found: {args.excel}", file=sys.stderr)
        return 2

    payload = build_payload(args.excel, parse_current_month(args.current_month))
    write_outputs(payload, args.out)
    print(
        f"Parsed {len(payload['workers'])} workers "
        f"across {len(payload['reporting_months'])} months → {args.out}"
    )
    for client in CLIENTS:
        slug = CLIENT_FILE_SLUG[client]
        n = len(json.loads((args.out / f"{slug}.json").read_text())["workers"])
        print(f"  {client:20s} → {slug}.json ({n} workers)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
