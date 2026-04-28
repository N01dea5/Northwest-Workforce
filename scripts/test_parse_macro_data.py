"""End-to-end sanity test for the Excel parser.

Builds a tiny in-memory `Rapidcrews Macro Data.xlsx` that mimics the sheets /
columns the real workbook has, runs the parser on it, and asserts the output
JSON shape.

Not framework-ed — just run `python3 scripts/test_parse_macro_data.py` and
it prints PASS / FAIL.
"""
from __future__ import annotations

import json
import shutil
import sys
import tempfile
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
from dateutil.relativedelta import relativedelta

SCRIPTS = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPTS))
import parse_macro_data as pm


def _build_fixture(tmp: Path, current_month: date) -> Path:
    wb = openpyxl.Workbook()
    # Personnel sheet ----------------------------------------------------
    ws = wb.active
    ws.title = "xll01 Personnel"
    ws.append(["Personnel Id", "Given Names", "Surname", "Primary Role", "Hire Company", "Start Date"])
    people = [
        ("P001", "Jane", "Smith", "Boilermaker", "SRG", date(2024, 1, 1)),
        ("P002", "John", "Doe", "Welder", "SRG", date(2025, 10, 1)),
        ("P003", "Mary", "O'Brien", "Advanced Rigger", "LabourHire", date(2023, 5, 1)),
        # No start date — parser should fall back to first_seen
        ("P004", "Bruce", "Wayne", "Scaffolder", "SRG", None),
    ]
    for row in people:
        ws.append(list(row))

    # Roster sheet -------------------------------------------------------
    roster = wb.create_sheet("xpbi02 PersonnelRosterView")
    roster.append([
        "Personnel Id", "Schedule Date", "Schedule Type",
        "Client", "Site", "Job No", "IsOnLocation",
    ])
    # Build a window: -3, -2, -1, current, +1, +2, +3 months
    months = [current_month + relativedelta(months=i) for i in range(-3, 4)]

    def shifts(pid, month_start, n, shift="Day Shift", client="Fortescue - Christmas Creek"):
        for i in range(n):
            d = month_start + timedelta(days=i)
            roster.append([pid, d, shift, client, "Site A", "J-001", True])

    # Jane: strong trailing utilisation at Fortescue, drops off in forward
    for m in months[:4]:
        shifts("P001", m, 14, "Day Shift", "Fortescue")
    # no forward work → at-risk candidate

    # John: steady Welder at BHP across window, including forward commits
    for m in months:
        shifts("P002", m, 12, "Day Shift", "BHP WAIO")

    # Mary: Rio Tinto all the way, heavy hours
    for m in months:
        shifts("P003", m, 13, "Day Shift", "Rio Tinto - Hamersley")

    # Bruce: Roy Hill, but some rows have IsOnLocation=False (should be skipped)
    for m in months:
        shifts("P004", m, 10, "Day Shift", "Roy Hill")
    # Add some "false" attendance rows — must be filtered out
    roster.append(["P004", months[0] + timedelta(days=20), "Day Shift", "Roy Hill", "Site", "J", False])

    # Row with non-NW client should be dropped
    roster.append(["P001", months[1] + timedelta(days=1), "Day Shift", "Covalent Mt Holland", "", "J", True])
    # RNR should contribute 0 hours but still count as rostered
    roster.append(["P002", months[2] + timedelta(days=5), "RNR", "BHP", "", "J", True])
    # Non-onsite schedule type → skip
    roster.append(["P001", months[1] + timedelta(days=2), "Personal Leave", "Fortescue", "", "J", True])

    # DailyPersonnelSchedule sheet — drives shutdown fulfilment.
    daily = wb.create_sheet("xpbi02 DailyPersonnelSchedule")
    daily.append([
        "PersonnelId", "ReportDate", "Status", "Trade",
        "JobId", "JobGroup", "JobStart", "Client",
    ])
    shutdown_start = months[1]
    # Jane: confirmed (filled & requested), then onsite (worked).
    daily.append(["P001", shutdown_start, "confirmed", "Boilermaker", "J100", "BHP Newman SD", shutdown_start, "BHP"])
    daily.append(["P001", shutdown_start + timedelta(days=1), "onsite", "Boilermaker", "J100", "BHP Newman SD", shutdown_start, "BHP"])
    # John: confirmed (filled & requested) but never onsite — counts as filled, not worked.
    daily.append(["P002", shutdown_start, "confirmed", "Welder", "J100", "BHP Newman SD", shutdown_start, "BHP"])
    # Mary: declined — counts as requested, not filled, declined outcome.
    daily.append(["P003", shutdown_start, "declined", "Advanced Rigger", "J100", "BHP Newman SD", shutdown_start, "BHP"])
    # Bruce: contacted — requested, not filled, no outcome.
    daily.append(["P004", shutdown_start, "contacted", "Scaffolder", "J100", "BHP Newman SD", shutdown_start, "BHP"])

    # ClientView — needed for JobPlanningView ClientId resolution.
    cv = wb.create_sheet("xpbi02 ClientView")
    cv.append(["ClientId", "ClientName"])
    cv.append(["BHP-CID", "BHP"])
    cv.append(["FMG-CID", "Fortescue"])
    cv.append(["RIO-CID", "RIO TINTO"])
    cv.append(["RH-CID",  "ROY HILL"])
    cv.append(["NOPE-CID", "Some Other Mob"])

    # DisciplineTrade — TradeId column drives JobPlanning competency lookup.
    dt = wb.create_sheet("xpbi02 DisciplineTrade")
    dt.append(["TradeId", "DisciplineId", "Discipline", "Trade"])
    dt.append(["TID-BOIL", "DID-1", "Mechanical", "Boilermaker"])
    dt.append(["TID-WELD", "DID-1", "Welder", "Coded Welder"])
    dt.append(["TID-RIG",  "DID-2", "Rigging", "Advanced Rigger"])

    # JobPlanningView — fulfilment rows.
    jp = wb.create_sheet("xpbi02 JobPlanningView")
    jp.append(["ClientId", "SiteId", "JobNo", "StartDate", "EndDate", "CompetencyId", "Required", "Filled", "ToFill", "ColourIndex", "JobId", "Actual"])
    in_window = months[1]      # one month behind current — inside reporting window
    out_window = months[0] + relativedelta(months=-6)  # well outside the window
    # BHP — boilermaker 5 req / 3 fill, welder 2 req / 2 fill (in window).
    jp.append(["BHP-CID", "S", 1, in_window, in_window, "TID-BOIL", 5, 3, 2, 0, "JID1", 0])
    jp.append(["BHP-CID", "S", 1, in_window, in_window, "TID-WELD", 2, 2, 0, 0, "JID1", 0])
    # Fortescue — rigger 4 req / 1 fill (in window).
    jp.append(["FMG-CID", "S", 2, in_window, in_window, "TID-RIG", 4, 1, 3, 0, "JID2", 0])
    # Out-of-window row — must be excluded.
    jp.append(["BHP-CID", "S", 3, out_window, out_window, "TID-BOIL", 99, 99, 0, 0, "JID3", 0])
    # Non-dashboard client — must be excluded.
    jp.append(["NOPE-CID", "S", 4, in_window, in_window, "TID-BOIL", 50, 50, 0, 0, "JID4", 0])

    excel = tmp / "Rapidcrews Macro Data.xlsx"
    wb.save(excel)
    return excel


def _run(current_month: date):
    tmp = Path(tempfile.mkdtemp(prefix="nw-parse-test-"))
    try:
        excel = _build_fixture(tmp, current_month)
        out = tmp / "data"
        rc = pm.main([
            "--excel", str(excel),
            "--out", str(out),
            "--current-month", current_month.strftime("%Y-%m"),
        ])
        assert rc == 0, f"parser returned {rc}"

        payload = json.loads((out / "workforce.json").read_text())
        assert payload["current_month"] == current_month.strftime("%Y-%m")
        assert len(payload["reporting_months"]) == 7
        assert payload["clients"] == ["Fortescue", "Rio Tinto", "BHP", "Roy Hill/Hancock"]

        workers = {w["name"]: w for w in payload["workers"]}
        assert set(workers) == {"Jane Smith", "John Doe", "Mary O'Brien", "Bruce Wayne"}, \
            f"unexpected names: {sorted(workers)}"

        # Jane — primary client Fortescue, trailing hours > 0, future months 0
        jane = workers["Jane Smith"]
        assert jane["primary_client"] == "Fortescue"
        cur_key = payload["current_month"]
        fwd_keys = payload["reporting_months"][payload["reporting_months"].index(cur_key) + 1:]
        for fk in fwd_keys:
            assert jane["monthly"][fk]["hours"] == 0, \
                f"expected Jane to drop off, saw {jane['monthly'][fk]}"
            assert jane["monthly"][fk]["committed"] is True

        # John — BHP, forward commitments present
        john = workers["John Doe"]
        assert john["primary_client"] == "BHP"
        assert john["monthly"][fwd_keys[0]]["hours"] > 0
        assert john["monthly"][fwd_keys[0]]["committed"] is True

        # Mary — Rio Tinto consistently
        mary = workers["Mary O'Brien"]
        assert mary["primary_client"] == "Rio Tinto"
        for mk in payload["reporting_months"]:
            assert mary["monthly"][mk]["client"] == "Rio Tinto", \
                f"Mary's client wrong at {mk}"

        # Bruce — Roy Hill, employment_start backfilled from first_seen
        bruce = workers["Bruce Wayne"]
        assert bruce["primary_client"] == "Roy Hill/Hancock"
        assert bruce["employment_start"].startswith(payload["reporting_months"][0])

        # Per-client files contain the expected subsets
        fort = json.loads((out / "fortescue.json").read_text())
        fort_names = {w["name"] for w in fort["workers"]}
        assert "Jane Smith" in fort_names
        assert "Mary O'Brien" not in fort_names, "Mary only worked Rio Tinto"

        rio = json.loads((out / "rio-tinto.json").read_text())
        rio_names = {w["name"] for w in rio["workers"]}
        assert rio_names == {"Mary O'Brien"}, f"rio names: {rio_names}"

        # Top positions should include the four we seeded
        for pos in ["Boilermaker", "Welder", "Advanced Rigger", "Scaffolder"]:
            assert pos in payload["positions_top20"], \
                f"missing {pos} in top20: {payload['positions_top20']}"

        # Non-NW client (Covalent) and "Personal Leave" rows must not show up
        # in Jane's Fortescue data — a "Personal Leave" entry shouldn't have
        # added any hours.
        # Jane's Fortescue-only hours should be ~ 14 days * 12h * 4 trailing months
        trailing = payload["reporting_months"][:4]
        jane_trailing_hours = sum(jane["monthly"][m]["hours"] for m in trailing)
        assert 500 <= jane_trailing_hours <= 800, \
            f"unexpected Jane trailing hours: {jane_trailing_hours}"

        # John's RNR contributed 0h (added on top of 12 day shifts that month)
        # so his hours for that month should be exactly 12 * 12.
        john_m2 = payload["reporting_months"][2]
        assert john["monthly"][john_m2]["hours"] == 12 * 12, \
            f"RNR should not add hours, saw {john['monthly'][john_m2]}"

        # Shutdown fulfilment: 4 trades requested, 2 filled (Jane + John),
        # Jane worked, Mary declined, Bruce no outcome.
        shutdowns = payload["shutdowns"]
        assert len(shutdowns) == 1, f"expected 1 shutdown, got {len(shutdowns)}"
        sd = shutdowns[0]
        assert sd["client"] == "BHP", f"shutdown client: {sd['client']}"
        assert sd["requested_total"] == 4, f"requested_total: {sd['requested_total']}"
        assert sd["filled_total"] == 2, f"filled_total: {sd['filled_total']}"
        trade_map = {t["trade"]: t for t in sd["trades"]}
        assert trade_map["Boilermaker"]["filled"] == 1
        assert trade_map["Welder"]["filled"] == 1
        assert trade_map["Advanced Rigger"]["filled"] == 0
        assert trade_map["Advanced Rigger"]["requested"] == 1
        assert trade_map["Scaffolder"]["filled"] == 0

        sid = sd["id"]
        assert jane["shutdown_outcomes"].get(sid) == "worked", \
            f"Jane outcome: {jane['shutdown_outcomes']}"
        assert mary["shutdown_outcomes"].get(sid) == "declined", \
            f"Mary outcome: {mary['shutdown_outcomes']}"
        assert bruce["shutdown_outcomes"].get(sid) is None, \
            f"Bruce outcome should be empty, got {bruce['shutdown_outcomes']}"
        # John is filled but never onsite — no worker outcome recorded.
        assert john["shutdown_outcomes"].get(sid) is None, \
            f"John outcome should be empty, got {john['shutdown_outcomes']}"

        # JobPlanningView fulfilment — clamped to dashboard clients +
        # reporting window. In-window BHP row: 5+2 req, 3+2 fill. In-window
        # Fortescue row: 4 req, 1 fill. Out-of-window and non-dashboard rows
        # must be dropped.
        ful = payload["fulfilment"]
        in_window_key = payload["reporting_months"][1]
        assert ful["totals"]["requested"] == 5 + 2 + 4
        assert ful["totals"]["filled"]    == 3 + 2 + 1
        bhp_row = ful["by_client_month"]["BHP"][in_window_key]
        assert bhp_row == {"requested": 7, "filled": 5}, bhp_row
        fmg_row = ful["by_client_month"]["Fortescue"][in_window_key]
        assert fmg_row == {"requested": 4, "filled": 1}, fmg_row
        # No Rio Tinto / Roy Hill rows in the fixture.
        assert ful["by_client_month"]["Rio Tinto"] == {}
        assert ful["by_client_month"]["Roy Hill/Hancock"] == {}
        # Trade rollup uses DisciplineTrade resolution.
        trades = {t["trade"]: t for t in ful["by_trade"]}
        assert trades["Boilermaker"]["requested"] == 5
        assert trades["Boilermaker"]["filled"]    == 3
        assert trades["Boilermaker"]["discipline"] == "Mechanical"
        assert trades["Advanced Rigger"]["requested"] == 4
        assert trades["Coded Welder"]["filled"] == 2

        print("PASS — parser produced expected workforce.json")
        return 0
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


if __name__ == "__main__":
    # Pin the fixture around a known month so the test is deterministic.
    rc = _run(date(2026, 4, 1))
    sys.exit(rc)
